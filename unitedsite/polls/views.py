import datetime
import random

import xlwt
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from django.shortcuts import render


from .models import Businessinfo, ClienteInfo
from .models import Userinfo
from django.conf import settings
from datetime import datetime
from datetime import timedelta


from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.base import View


# Create your views here.

def toLogin_view(request):
    return render(request, "login.html")


def login_view(request):
    u = request.POST.get("user", "")
    p = request.POST.get("pwd", "")
    if u and p:
        c = Userinfo.objects.filter(user_name=u, user_pwd=p).count()
        if c == 1:
            businessinfos = Businessinfo.objects.order_by('id')
            return render(request, "main.html", {"businessLists": businessinfos})
        else:
            return HttpResponse("用户名或密码错误")
    else:
        return HttpResponse("请输入用户名和密码")


def toregister_view(request):
    return render(request, "registry.html")


def register_view(request):
    u = request.POST.get("user", "")
    p = request.POST.get("pwd", "")
    if u and p:
        user = Userinfo(id=str(random.randrange(0000, 9999)), user_name=u, user_pwd=p)
        user.save()
        return HttpResponse("注册成功")
    else:
        return HttpResponse("请输入用户名和密码")


#def email_check(user):
#    return user.email.endswith('@unitedlogdom.com')
#@user_passes_test(email_check)
@login_required()
def main_view(request):
    businessinfos = Businessinfo.objects.filter(operation_status="NO").order_by("id")
    return render(request, "main.html", {"businessLists": businessinfos})

def new_business(request):
    userSelectUID= request.POST.get("userSelectBID", "")
    if userSelectUID == "": #新建
        return render(request, "newbusiness.html")
    else: # 修改
        businessinfo = Businessinfo.objects.get(id=userSelectUID)
        return render(request, "modbusiness.html", {"businessinfo": businessinfo})

def savebusiness(request):
    id = request.POST.get("id", "")
    clientname = request.POST.get("clientname", "")
    week = request.POST.get("week", "")
    estimated_load_date = request.POST.get("estimated_load_date", None)
    if estimated_load_date=="":
        estimated_load_date=None
    pre_etd = request.POST.get("pre_etd", None)
    if pre_etd=="":
        pre_etd=None
    etd = request.POST.get("etd", None)
    if etd == "":
        etd = None
    eta = request.POST.get("eta", None)
    if eta == "":
        eta = None
    vessel_voyage = request.POST.get("vessel_voyage", "")
    pol = request.POST.get("pol", "")
    pod = request.POST.get("pod", "")
    businessinfocol = request.POST.get("businessinfocol", "")
    numctrs = request.POST.get("numctrs", "")
    type_ctrs = request.POST.get("type_ctrs", "")
    booking_no = request.POST.get("booking_no", "")
    master_no = request.POST.get("master_no", "")
    first_hbl_no = request.POST.get("first_hbl_no", "")
    hbl_memo = request.POST.get("hbl_memo", "")
    num_hbls = request.POST.get("num_hbls", None)
    if num_hbls == "":
        num_hbls = None
    container_no = request.POST.get("container_no", "")
    shipping_line = request.POST.get("shipping_line", "")
    freeddday = request.POST.get("freeddday", None)
    cost_org = request.POST.get("cost_org", None)
    if cost_org == "":
        cost_org = None
    sale_org = request.POST.get("sale_org", None)
    if sale_org == "":
        sale_org = None
    set_org = request.POST.get("set_org", None)
    if set_org == "":
        set_org = None
    local_org = request.POST.get("local_org", None)
    if local_org == "":
        local_org = None
    venta_dest = request.POST.get("venta_dest", None)
    if venta_dest == "":
        venta_dest = None
    costhbl_dest_description = request.POST.get("costhbl_dest_description", "")
    costhbl_dest = request.POST.get("costhbl_dest", None)
    if costhbl_dest == "":
        costhbl_dest = None
    costhbl_dest_status = request.POST.get("costhbl_dest_status", "NO")
    topaid_org_description = request.POST.get("topaid_org_description", "")
    topaid_org = request.POST.get("topaid_org", None)
    if topaid_org == "":
        topaid_org = None
    topaid_status = request.POST.get("topaid_status", "NO")
    toinvoice_dest_description = request.POST.get("toinvoice_dest_description", "")
    toinvoice_dest = request.POST.get("toinvoice_dest", None)
    if toinvoice_dest == "":
        toinvoice_dest = None
    toinvoice_dest_status = request.POST.get("toinvoice_dest_status", "NO")
    profit = request.POST.get("profit", None)
    if profit == "":
        profit = None
    hbl_canjeado_status = request.POST.get("hbl_canjeado_status", "NO")
    devolution_ctrs_inidate = request.POST.get("devolution_ctrs_inidate", None)
    if devolution_ctrs_inidate == "":
        devolution_ctrs_inidate = None
    devolution_ctrs_findate = request.POST.get("devolution_ctrs_findate", None)
    if devolution_ctrs_findate == "":
        devolution_ctrs_findate = None
    operation_status = request.POST.get("operation_status", "NO")
    bl_type = request.POST.get("bl_type", "")
    release_type = request.POST.get("release_type", "")
    mblfile = request.FILES.get("mblfile", "")
    hblfiles = request.FILES.get("hblfiles", "")
    dp_voucher = request.FILES.get("dp_voucher", "")
    dp_invoice = request.FILES.get("dp_invoice", "")
    hp_voucher = request.FILES.get("hp_voucher", "")
    hp_invoice = request.FILES.get("hp_invoice", "")
    hbls_canjeados = request.FILES.get("hbls_canjeados", "")
    income_voucher = request.FILES.get("income_voucher", "")
    income_invoice = request.FILES.get("income_invoice", "")
    hbls_firmados = request.FILES.get("hbls_firmados", "")
    if vessel_voyage and pol and pod and numctrs and type_ctrs and booking_no and shipping_line and freeddday and \
    costhbl_dest_status and topaid_status and toinvoice_dest_status and hbl_canjeado_status and operation_status:
        if id: # 修改保存
            businessinfo = Businessinfo.objects.get(id = id)
            businessinfo.clientname=clientname
            businessinfo.week=week
            businessinfo.estimated_load_date=estimated_load_date
            businessinfo.pre_etd=pre_etd
            businessinfo.etd=etd
            businessinfo.eta=eta
            businessinfo.vessel_voyage=vessel_voyage
            businessinfo.pol=pol
            businessinfo.pod=pod
            businessinfo.businessinfocol=businessinfocol
            businessinfo.numctrs=numctrs
            businessinfo.type_ctrs=type_ctrs
            businessinfo.booking_no=booking_no
            businessinfo.master_no=master_no
            businessinfo.first_hbl_no = first_hbl_no
            businessinfo.hbl_memo = hbl_memo
            businessinfo.num_hbls=num_hbls
            businessinfo.container_no=container_no
            businessinfo.shipping_line=shipping_line
            businessinfo.freeddday=freeddday
            businessinfo.cost_org=cost_org
            businessinfo.sale_org=sale_org
            businessinfo.set_org=set_org
            businessinfo.local_org=local_org
            businessinfo.venta_dest=venta_dest
            businessinfo.costhbl_dest_description=costhbl_dest_description
            businessinfo.costhbl_dest=costhbl_dest
            businessinfo.costhbl_dest_status=costhbl_dest_status
            businessinfo.topaid_org_description=topaid_org_description
            businessinfo.topaid_org=topaid_org
            businessinfo.topaid_status=topaid_status
            businessinfo.toinvoice_dest_description=toinvoice_dest_description
            businessinfo.toinvoice_dest=toinvoice_dest
            businessinfo.toinvoice_dest_status=toinvoice_dest_status
            businessinfo.profit=profit
            businessinfo.hbl_canjeado_status=hbl_canjeado_status
            businessinfo.devolution_ctrs_inidate=devolution_ctrs_inidate
            businessinfo.devolution_ctrs_findate=devolution_ctrs_findate
            businessinfo.operation_status=operation_status
            businessinfo.bl_type=bl_type
            businessinfo.release_type=release_type
            if mblfile:
                businessinfo.mblfile = mblfile
            if hblfiles:
                businessinfo.hblfiles = hblfiles
            if dp_voucher:
                businessinfo.dp_voucher = dp_voucher
            if dp_invoice:
                businessinfo.dp_invoice = dp_invoice
            if hp_voucher:
                businessinfo.hp_voucher = hp_voucher
            if hp_invoice:
                businessinfo.hp_invoice = hp_invoice
            if hbls_canjeados:
                businessinfo.hbls_canjeados = hbls_canjeados
            if income_voucher:
                businessinfo.income_voucher = income_voucher
            if income_invoice:
                businessinfo.income_invoice = income_invoice
            if hbls_firmados:
                businessinfo.hbls_firmados = hbls_firmados


        else: # 新增保存
            id = Businessinfo.objects.all().order_by("-id")[0].id+1
            businessinfo = Businessinfo( \
                id = id,
                clientname=clientname,
                week=week,
                estimated_load_date=estimated_load_date,
                pre_etd=pre_etd,
                etd=etd,
                eta=eta,
                vessel_voyage=vessel_voyage,
                pol=pol,
                pod=pod,
                businessinfocol=businessinfocol,
                numctrs=numctrs,
                type_ctrs=type_ctrs,
                booking_no=booking_no,
                master_no=master_no,
                first_hbl_no = first_hbl_no,
                hbl_memo = hbl_memo,
                num_hbls=num_hbls,
                container_no=container_no,
                shipping_line=shipping_line,
                freeddday=freeddday,
                cost_org=cost_org,
                sale_org=sale_org,
                set_org=set_org,
                local_org=local_org,
                venta_dest=venta_dest,
                costhbl_dest_description=costhbl_dest_description,
                costhbl_dest=costhbl_dest,
                costhbl_dest_status=costhbl_dest_status,
                topaid_org_description=topaid_org_description,
                topaid_org=topaid_org,
                topaid_status=topaid_status,
                toinvoice_dest_description=toinvoice_dest_description,
                toinvoice_dest=toinvoice_dest,
                toinvoice_dest_status=toinvoice_dest_status,
                profit=profit,
                hbl_canjeado_status=hbl_canjeado_status,
                devolution_ctrs_inidate=devolution_ctrs_inidate,
                devolution_ctrs_findate=devolution_ctrs_findate,
                operation_status=operation_status,
                bl_type=bl_type,
                release_type=release_type,
                mblfile=mblfile,
                hblfiles=hblfiles,
                dp_voucher=dp_voucher,
                dp_invoice=dp_invoice,
                hp_voucher=hp_voucher,
                hp_invoice=hp_invoice,
                hbls_canjeados=hbls_canjeados,
                income_voucher = income_voucher,
                income_invoice = income_invoice,
                hbls_firmados = hbls_firmados,
            )

        businessinfo.save()
        businessinfos = Businessinfo.objects.filter(operation_status="NO").order_by("id")
        return render(request, "main.html", {"businessLists": businessinfos})
    else: # 必填数据为空
        return HttpResponse("请输入数据")

def delbusiness(request):
    idbusinessfordelete = request.POST.get("idbusinessfordelete", "")
    if idbusinessfordelete:
        businessinfo = Businessinfo.objects.get(id=idbusinessfordelete)
        businessinfo.delete()
        businessinfos = Businessinfo.objects.filter(operation_status="NO").order_by("id")
        return render(request, "main.html", {"businessLists": businessinfos})

def prealerta(request):
    idbusinessforprealerta = request.POST.get("idbusinessforprealerta", "")
    businessinfo = Businessinfo.objects.get(id=idbusinessforprealerta)
    clienteInfo = ClienteInfo.objects.get(id=businessinfo.clientname)
    clientename=clienteInfo.clientname
    container_no=Businessinfo.objects.get(id=idbusinessforprealerta).container_no;


    deta=datetime.strptime(str(businessinfo.eta),"%Y-%m-%d")
    deta1=timedelta(days=-14)
    dateline=deta+deta1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=PREALERT_BL_'+ str(businessinfo.booking_no)+ "_"+ container_no +"_"+clientename + "_"+ str(datetime.today().month)+'.'+str(datetime.today().day)+'.xlsx'

    wb = load_workbook(filename="."+settings.STATIC_URL+"Template/tPreAlerta.xlsx") #, read_only=True
    sheet = wb.get_sheet_by_name("PRE-PAYMENT-ALERT")
    img= Image("."+settings.STATIC_URL+"img/logo.png")
    newsize=(145,32)
    img.width, img.height = newsize
    sheet.add_image(img,"B4")
    sheet["C12"] = clienteInfo.clientname
    sheet["C13"] = clienteInfo.clientrut
    sheet["C14"] = clienteInfo.clientgiro
    sheet["C15"] = clienteInfo.clientaddress
    sheet["C16"] = clienteInfo.clientstate
    sheet["F16"] = clienteInfo.clientcity
    sheet["C17"] = clienteInfo.clientcontact
    sheet["I12"] = str(datetime.today().date())

    sheet["D22"] ="Flete Maritimo "+ businessinfo.shipping_line+ " Line " + str(businessinfo.numctrs) +"X"+ businessinfo.type_ctrs
    pricesList=businessinfo.toinvoice_dest_description.split("+")
    i=0
    totalprice = 0
    for i in range(len(pricesList)):
        if (i==0):
            sheet["B23"] = businessinfo.numctrs
            sheet["C23"] = "Ctr"
            sheet["D23"] = "BL " + businessinfo.booking_no + " / CTR: " + businessinfo.container_no
            sheet["D24"] = businessinfo.vessel_voyage
            sheet["D25"] = businessinfo.pol + " - " + businessinfo.pod
            sheet["H23"] = float(pricesList[0])
            sheet["I23"] = sheet["H23"].value
            totalprice +=float(pricesList[0])
        elif (i==1):
            sheet["B27"] = businessinfo.numctrs
            sheet["C27"] = "Ctr"
            sheet["D27"] = "Logistics at Destination:"
            sheet["D28"] = "Handling + Apertura + Emision BL + Carta de Responsabilidad"
            sheet["H27"] = float(pricesList[1])
            sheet["I27"] = sheet["H27"].value
            totalprice += float(pricesList[1])
        elif (i == 2):
            sheet["B29"] = businessinfo.numctrs
            sheet["C29"] = "Ctr"
            sheet["D29"] = "DTHC"
            sheet["H29"] = float(pricesList[2])
            sheet["I29"] = sheet["H29"].value
            totalprice += float(pricesList[2])
        elif (i == 3):
            sheet["B30"] = businessinfo.numctrs
            sheet["C30"] = "Ctr"
            sheet["D30"] = "EXTRA HBL/ORIGINAL FEE/GATE IN FEE"
            sheet["H30"] = float(pricesList[3])
            sheet["I30"] = sheet["H30"].value
            totalprice += float(pricesList[3])
        else:
            print("Have more prices...")
            totalprice += float(pricesList[i])


        i += 1

    sheet["D32"]="Amount payable US"+ '${:,.0f}'.format(totalprice)
    sheet["D33"] = "Please pay our dollar account below"

    sheet["D35"] = "Payment should be executed before " + str(dateline.date())

    sheet["B39"] = "美元(大写)金额: "
    sheet["D39"] = totalprice

    sheet["I38"] = totalprice
    sheet["I42"] = totalprice

    sheet["C50"] = ""
    sheet["C51"] = ""
    sheet["C52"] = ""

    sheet["F50"] = ""



    #h1= sheet.cell(row=6 ,column=8).value

    h1 = sheet.cell(row=13, column=3).value
    print("#Prealterta:"+str(h1))

    wb.save(response)
    return response
    #return HttpResponse("PreAlerta generado!")

def exportbusiness(request):
    print(request.POST.get("smtExpOrigen",""))
    print(request.POST.get("smtExpCompany",""))
    response = HttpResponse(content_type='application/ms-excel')
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("Export data")
    row_num = 0

    # 第一行格式
    titleStyle = xlwt.Style.easyxf('pattern:pattern solid, fore_colour dark_blue;'
                                   'align: vertical center, horizontal center;'
                                   'font: bold true, colour white;'
                                   )
    dataStyle = xlwt.Style.easyxf('pattern:pattern solid, fore_colour white;'
                                  'align: vertical center, horizontal center;'
                                  'font: bold false, colour black;')

    finishStyle = xlwt.Style.easyxf('pattern:pattern solid, fore_colour gray40;'
                                    'align: vertical center, horizontal center;'
                                    'font: bold false, colour black;')

    titleStyle.font.name = "Times New Roman"
    titleStyle.font.height = 220

    dataStyle.font.name = "Times New Roman"
    dataStyle.font.height = 180

    finishStyle.font.name = "Times New Roman"
    finishStyle.font.height = 180

    borders = xlwt.Borders()
    borders.left = 1
    borders.right = 1
    borders.top = 1
    borders.bottom = 1
    titleStyle.borders = borders
    dataStyle.borders = borders
    finishStyle.borders = borders




    if request.POST.get("smtExpOrigen")!=None: # 给客户公司
        response['Content-Disposition'] = 'attachment; filename='+str(datetime.today().month)+'.'+\
                                      str(datetime.today().day)+'Chile_Business_Table(Update_by_Chile).xls'
        print("文件路径：" + str(datetime.today().month)+'.'+str(datetime.today().day)+
              'Chile_Business_Table(Update_by_Chile).xls')

        columns = ['ID', '客户', 'WEEK', 'ETD', '实际ETD',
                   '预计装箱日期', '船名航次', '起始港', '目的港',
                   '箱量', '箱型', '订舱号', '箱号', '船司', '免箱期',
                   '运价成本', '卖价', '贵司成本', '付款状态', '备注信息 ']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], titleStyle)

        rows = Businessinfo.objects.values_list('id',
                                                'clientname',
                                                'week',
                                                'pre_etd',
                                                'etd',
                                                'estimated_load_date',
                                                'vessel_voyage',
                                                'pol',
                                                'pod',
                                                'numctrs',
                                                'type_ctrs',
                                                'booking_no',
                                                'container_no',
                                                'shipping_line',
                                                'freeddday',
                                                'cost_org',
                                                'sale_org',
                                                'set_org',
                                                'topaid_status',
                                                'businessinfocol')

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):

                if col_num == 0:
                    ws.col(col_num).width = 1000
                    dataStyle.num_format_str = "0"
                    finishStyle.num_format_str = "0"
                elif col_num == 1:
                    ws.col(col_num).width = 5000
                elif col_num == 2:
                    ws.col(col_num).width = 2100
                elif col_num == 3:
                    ws.col(col_num).width = 2300
                    dataStyle.num_format_str = "YYYY-MM-DD"
                    finishStyle.num_format_str = "YYYY-MM-DD"
                elif col_num == 4:
                    ws.col(col_num).width = 2500
                    dataStyle.num_format_str = "YYYY-MM-DD"
                    finishStyle.num_format_str = "YYYY-MM-DD"
                elif col_num == 5:
                    ws.col(col_num).width = 3800
                    dataStyle.num_format_str = "YYYY-MM-DD"
                    finishStyle.num_format_str = "YYYY-MM-DD"
                elif col_num == 6:
                    ws.col(col_num).width = 5800
                elif col_num == 7:
                    ws.col(col_num).width = 3100
                elif col_num == 8:
                    ws.col(col_num).width = 3500
                elif col_num == 9:
                    ws.col(col_num).width = 1000
                    dataStyle.num_format_str = "0"
                    finishStyle.num_format_str = "0"
                elif col_num == 10:
                    ws.col(col_num).width = 1300
                elif col_num == 11:
                    ws.col(col_num).width = 4300
                elif col_num == 12:
                    ws.col(col_num).width = 3400
                elif col_num == 13:
                    ws.col(col_num).width = 1900
                elif col_num == 14:
                    ws.col(col_num).width = 2000
                    dataStyle.num_format_str = "0"
                elif col_num == 15:
                    ws.col(col_num).width = 3500
                    dataStyle.num_format_str = '[$US$] #,##0'
                    finishStyle.num_format_str = "[$US$] #,##0"
                elif col_num == 16:
                    ws.col(col_num).width = 3500
                    dataStyle.num_format_str = '[$US$] #,##0'
                    finishStyle.num_format_str = "[$US$] #,##0"
                elif col_num == 17:
                    ws.col(col_num).width = 3500
                    dataStyle.num_format_str = '[$US$] #,##0'
                    finishStyle.num_format_str = "[$US$] #,##0"
                elif col_num == 18:
                    ws.col(col_num).width = 2600
                else:
                    ws.col(col_num).width = 5000
                if row[18] == "NO":
                    ws.write(row_num, col_num, row[col_num], dataStyle)
                else:
                    ws.write(row_num, col_num, row[col_num], finishStyle)

    elif request.POST.get("smtExpCompany")!=None: # 自己公司
        titleStyle.font.height = 180
        response['Content-Disposition'] = 'attachment; filename=' + str(datetime.today().month) + '.' + \
                                          str(datetime.today().day) + 'Logistic_Chile_Contacto.xls'
        print("文件路径：" + str(datetime.today().month) + '.' + str(datetime.today().day) +
              'Logistic_Chile_Contacto.xls')

        columns = ['ID#', 'Cliente', 'ETA(POD)','Vessel Voyage','POL','POD','C.Ctrs', 'T.Ctr',
                   'Master#', 'HBL#(Booking#)','HBL Memo','C.HBLs', 'CTR#', 'Naviera',
                   'Costo Chile Detalle', 'Costo Chile', 'Costo China Detalle','Costo China',
                   'Venta Detalle', 'Venta','Margen Bruto','Costo Chile Pagado?','Costo China Pagado?',
                   'Venta Recibido?','HBL Canjeado?','Tipo BL','Tipo Liberacion','Fecha Inicio Dvltn','Fecha Fin Dvltn','D&D Libre','Estado de Cierra (todo OK?)']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], titleStyle)

        rows = Businessinfo.objects.values_list('id',
                                                'clientname',
                                                'eta',
                                                'vessel_voyage',
                                                'pol',
                                                'pod',
                                                'numctrs',
                                                'type_ctrs',
                                                'master_no',
                                                'first_hbl_no',
                                                'hbl_memo',
                                                'num_hbls',
                                                'container_no',
                                                'shipping_line',

                                                'costhbl_dest_description',
                                                'costhbl_dest',
                                                'topaid_org_description',
                                                'topaid_org',
                                                'toinvoice_dest_description',
                                                'toinvoice_dest',
                                                'profit',
                                                'costhbl_dest_status',
                                                'topaid_status',
                                                'toinvoice_dest_status',
                                                'hbl_canjeado_status',
                                                'bl_type',
                                                'release_type',
                                                'devolution_ctrs_inidate',
                                                'devolution_ctrs_findate',
                                                'freeddday',
                                                'operation_status')

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if col_num == 0:
                    ws.col(col_num).width = 1000
                    dataStyle.num_format_str = "0"
                    finishStyle.num_format_str = "0"
                elif col_num == 1:
                    ws.col(col_num).width = 5000
                elif col_num == 2:
                    ws.col(col_num).width = 2300
                    dataStyle.num_format_str = "YYYY-MM-DD"
                    finishStyle.num_format_str = "YYYY-MM-DD"
                elif col_num == 3:
                    ws.col(col_num).width = 5800
                elif col_num == 4:
                    ws.col(col_num).width = 3100
                elif col_num == 5:
                    ws.col(col_num).width = 3500
                elif col_num == 6:
                    ws.col(col_num).width = 1000
                    dataStyle.num_format_str = "0"
                    finishStyle.num_format_str = "0"
                elif col_num == 7:
                    ws.col(col_num).width = 1300
                elif col_num == 8:
                    ws.col(col_num).width = 4300
                elif col_num == 9:
                    ws.col(col_num).width = 4300
                elif col_num == 10:
                    ws.col(col_num).width = 4300
                elif col_num == 11:
                    ws.col(col_num).width = 1500
                elif col_num == 12:
                    ws.col(col_num).width = 3400
                elif col_num == 13:
                    ws.col(col_num).width = 1900

                elif col_num == 14:
                    ws.col(col_num).width = 4000
                elif col_num == 15:
                    ws.col(col_num).width = 3500
                    dataStyle.num_format_str = '[$US$] #,##0'
                    finishStyle.num_format_str = "[$US$] #,##0"
                elif col_num == 16:
                    ws.col(col_num).width = 4000
                elif col_num == 17:
                    ws.col(col_num).width = 3500
                    dataStyle.num_format_str = '[$US$] #,##0'
                    finishStyle.num_format_str = "[$US$] #,##0"
                elif col_num == 18:
                    ws.col(col_num).width = 4000
                elif col_num == 19:
                    ws.col(col_num).width = 3500
                    dataStyle.num_format_str = '[$US$] #,##0'
                    finishStyle.num_format_str = "[$US$] #,##0"
                elif col_num == 20:
                    ws.col(col_num).width = 3500
                    dataStyle.num_format_str = '[$US$] #,##0'
                    finishStyle.num_format_str = "[$US$] #,##0"
                elif col_num == 21:
                    ws.col(col_num).width = 2600
                elif col_num == 22:
                    ws.col(col_num).width = 2600
                elif col_num == 23:
                    ws.col(col_num).width = 2600
                elif col_num == 24:
                    ws.col(col_num).width = 2600

                elif col_num == 25:
                    ws.col(col_num).width = 2500
                elif col_num == 26:
                    ws.col(col_num).width = 5000

                elif col_num == 27:
                    ws.col(col_num).width = 2500
                    dataStyle.num_format_str = "YYYY-MM-DD"
                    finishStyle.num_format_str = "YYYY-MM-DD"
                elif col_num == 28:
                    ws.col(col_num).width = 2500
                    dataStyle.num_format_str = "YYYY-MM-DD"
                    finishStyle.num_format_str = "YYYY-MM-DD"
                elif col_num == 29:
                    ws.col(col_num).width = 2000
                    dataStyle.num_format_str = "0"
                    finishStyle.num_format_str = "0"
                elif col_num == 30:
                    ws.col(col_num).width = 2600
                else:
                    ws.col(col_num).width = 5000
                if row[30] == "NO":
                    ws.write(row_num, col_num, row[col_num], dataStyle)
                else:
                    ws.write(row_num, col_num, row[col_num], finishStyle)

    wb.save(response)
    return response

def search(request):

    finish_estado = request.POST.get("finish_estado","NO")
    cond_Cliente = request.POST.get("cond_Cliente", "")

    print("estado:"+finish_estado)
    print("cond_Cliente:" + cond_Cliente)




    if finish_estado=="YES" and cond_Cliente!="":
        businessinfos = Businessinfo.objects.filter(clientname=cond_Cliente).order_by("id")


    elif finish_estado=="YES" and cond_Cliente=="":
        businessinfos = Businessinfo.objects.filter().order_by("id")
    elif cond_Cliente!="" and finish_estado=="NO":
        businessinfos = Businessinfo.objects.filter(operation_status=finish_estado,clientname=cond_Cliente).order_by("id")
    else:
        businessinfos = Businessinfo.objects.filter(operation_status="NO").order_by("id")
    return render(request, "main.html", {"businessLists": businessinfos, "finish_estado": finish_estado, "cond_Cliente": cond_Cliente})


def importbusiness(request):
    return HttpResponse("ok")

def profile(request):
    return HttpResponse("ok")

# 此类中的函数只能在登录后使用
'''
class AuthedView(LoginRequiredMixin,View):
    login_url = '/accounts/login/'
    #测试用户每个权限
    def test_func(self):
        return self.request.user.email.endswith('@unitedlogdom.com')
    def profile(request):
        return return HttpResponse("ok")
'''